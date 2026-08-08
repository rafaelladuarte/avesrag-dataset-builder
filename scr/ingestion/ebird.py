import logging
import time
import unicodedata
from pathlib import Path
from typing import Optional

import requests
from openpyxl import load_workbook
from pymongo import MongoClient, UpdateOne
from pymongo.errors import BulkWriteError
from rapidfuzz import fuzz, process
from tqdm import tqdm

from scr.core.config import settings

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------

EBIRD_API_KEY: str = settings.EBIRD_API_KEY or ""
MONGO_URI: str = str(settings.MONGODB_URI)
MONGO_DB: str = settings.MONGODB_DATABASE

# Caminho para o xlsx do IBGE — ajuste se necessário
IBGE_XLSX: Path = Path("data/raw/Bioma_Predominante_por_Municipio_2024.xlsx")

# Coluna exata do xlsx (case-sensitive). Ajuste se os cabeçalhos forem diferentes.
COL_GEOCODIGO = "Geocódigo"
COL_NOME = "Nome do município"
COL_UF = "Sigla da UF"
COL_BIOMA = "Bioma predominante"

# Intervalo entre chamadas à API (segundos) — respeite o rate limit do eBird
SLEEP_BETWEEN_REQUESTS: float = 0.6

# Score mínimo de similaridade para o fuzzy match (0–100)
FUZZY_THRESHOLD: int = 85

# Locale para nomes populares das espécies
TAXONOMY_LOCALE: str = "pt"

BASE_URL = "https://api.ebird.org/v2"

# Handler para o arquivo de log (sempre ativo)
_file_handler = logging.FileHandler("ebird_coleta.log", encoding="utf-8")
_file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

# Handler para o console — usamos tqdm.write nos pontos críticos para não
# quebrar a barra de progresso; aqui só capturamos o restante.
_console_handler = logging.StreamHandler()
_console_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[_console_handler, _file_handler],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def normalize(text: str) -> str:
    """Lowercase, remove acentos e apóstrofos para comparação."""
    text = text.strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.replace("'", "").replace("`", "")
    return text


def ebird_get(path: str, params: Optional[dict] = None) -> list | dict:
    """GET autenticado na API eBird com retry e backoff exponencial."""
    headers = {"X-eBirdApiToken": EBIRD_API_KEY}
    url = f"{BASE_URL}{path}"
    for attempt in range(4):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=30)
            r.raise_for_status()
            return r.json()
        except requests.HTTPError:
            if r.status_code == 429:
                # Backoff exponencial: 10s → 20s → 40s → 80s
                wait = 10 * (2**attempt)
                msg = f"Rate limit atingido (tentativa {attempt + 1}/4). Aguardando {wait}s…"
                # Usa tqdm.write para não corromper a barra de progresso no console
                tqdm.write(f"[WARNING] {msg}")
                # Registra também no arquivo de log sem duplicar no console
                _file_handler.emit(
                    logging.makeLogRecord(
                        {
                            "levelno": logging.WARNING,
                            "levelname": "WARNING",
                            "msg": msg,
                            "args": (),
                            "name": __name__,
                        }
                    )
                )
                time.sleep(wait)
            else:
                raise
        except requests.RequestException as e:
            err_msg = f"Erro na requisição {url}: {e}. Tentativa {attempt + 1}/4"
            tqdm.write(f"[ERROR] {err_msg}")
            _file_handler.emit(
                logging.makeLogRecord(
                    {
                        "levelno": logging.ERROR,
                        "levelname": "ERROR",
                        "msg": err_msg,
                        "args": (),
                        "name": __name__,
                    }
                )
            )
            time.sleep(3)
    raise RuntimeError(f"Falha ao acessar {url} após 4 tentativas")


# ---------------------------------------------------------------------------
# Etapa 1 — Ler base IBGE
# ---------------------------------------------------------------------------


def load_ibge(path: Path) -> list[dict]:
    """Carrega o xlsx do IBGE e retorna lista de dicts."""
    log.info(f"Lendo {path}…")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [
        str(cell.value).strip() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=2, max_row=2))
    ]
    idx = {h: i for i, h in enumerate(headers)}

    required = [COL_GEOCODIGO, COL_NOME, COL_UF, COL_BIOMA]
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(
            f"Colunas não encontradas no xlsx: {missing}\nColunas disponíveis: {headers}"
        )

    municipios = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        geocodigo = str(row[idx[COL_GEOCODIGO]]).strip() if row[idx[COL_GEOCODIGO]] else None
        nome = str(row[idx[COL_NOME]]).strip() if row[idx[COL_NOME]] else None
        uf = str(row[idx[COL_UF]]).strip() if row[idx[COL_UF]] else None
        bioma = str(row[idx[COL_BIOMA]]).strip() if row[idx[COL_BIOMA]] else None

        if nome and uf:
            municipios.append(
                {
                    "geocodigo": geocodigo,
                    "nome": nome,
                    "uf": uf,
                    "bioma": bioma,
                    "nome_norm": normalize(nome),
                }
            )

    wb.close()
    log.info(f"  → {len(municipios)} municípios carregados")
    return municipios


# ---------------------------------------------------------------------------
# Etapa 2 — Buscar municípios do eBird por UF e cruzar com IBGE
# ---------------------------------------------------------------------------


def fetch_ebird_municipios(uf: str) -> list[dict]:
    """Retorna lista [{code, name, name_norm}] para a UF dada."""
    data = ebird_get(f"/ref/region/list/subnational2/BR-{uf}")
    time.sleep(SLEEP_BETWEEN_REQUESTS)
    return [{"code": r["code"], "name": r["name"], "name_norm": normalize(r["name"])} for r in data]


def crossmatch(ibge_list: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Cruza ibge_list com eBird município a município.
    Retorna (matched, unmatched).
    """
    ufs = sorted({m["uf"] for m in ibge_list})
    matched, unmatched = [], []

    log.info("Cruzando municípios IBGE ↔ eBird…")
    for uf in tqdm(ufs, desc="UFs"):
        try:
            ebird_muns = fetch_ebird_municipios(uf)
        except Exception as e:
            log.error(f"Erro ao buscar municípios de {uf}: {e}")
            continue

        ebird_norms = [m["name_norm"] for m in ebird_muns]
        ebird_by_norm = {m["name_norm"]: m for m in ebird_muns}

        ibge_uf = [m for m in ibge_list if m["uf"] == uf]

        for mun in ibge_uf:
            norm = mun["nome_norm"]

            # 1. Tentativa exata
            if norm in ebird_by_norm:
                mun["ebird_code"] = ebird_by_norm[norm]["code"]
                mun["ebird_name"] = ebird_by_norm[norm]["name"]
                mun["match_type"] = "exact"
                matched.append(mun)
                continue

            # 2. Fuzzy fallback
            result = process.extractOne(norm, ebird_norms, scorer=fuzz.token_sort_ratio)
            if result and result[1] >= FUZZY_THRESHOLD:
                best_norm = result[0]
                mun["ebird_code"] = ebird_by_norm[best_norm]["code"]
                mun["ebird_name"] = ebird_by_norm[best_norm]["name"]
                mun["match_type"] = f"fuzzy:{result[1]}"
                matched.append(mun)
            else:
                mun["ebird_code"] = None
                mun["match_type"] = "no_match"
                unmatched.append(mun)

    log.info(f"  → matched: {len(matched)} | sem match: {len(unmatched)}")
    return matched, unmatched


# ---------------------------------------------------------------------------
# Etapa 3 — Taxonomia eBird (cache único)
# ---------------------------------------------------------------------------


def fetch_taxonomy(locale: str = TAXONOMY_LOCALE) -> dict[str, dict]:
    """Baixa a taxonomia completa e retorna dict keyed by speciesCode."""
    log.info("Baixando taxonomia eBird (cache único)…")
    data = ebird_get("/ref/taxonomy/ebird", params={"fmt": "json", "locale": locale})
    taxonomy = {}
    for t in data:
        taxonomy[t["speciesCode"]] = {
            "speciesCode": t["speciesCode"],
            "comName": t.get("comName", ""),
            "sciName": t.get("sciName", ""),
            "order": t.get("order", ""),
            "familyComName": t.get("familyComName", ""),
            "familySciName": t.get("familySciName", ""),
            "category": t.get("category", ""),
        }
    log.info(f"  → {len(taxonomy)} táxons carregados")
    return taxonomy


# ---------------------------------------------------------------------------
# Etapa 4 — Buscar espécies por município
# ---------------------------------------------------------------------------


def fetch_species_for_municipio(ebird_code: str, taxonomy: dict) -> list[dict]:
    """
    Chama /product/spplist/{regionCode} e enriquece com taxonomia.
    Retorna lista de dicts de espécie.
    """
    try:
        codes = ebird_get(f"/product/spplist/{ebird_code}")
        time.sleep(SLEEP_BETWEEN_REQUESTS)
    except Exception as e:
        log.warning(f"  Sem dados para {ebird_code}: {e}")
        return []

    species = []
    for code in codes:
        info = taxonomy.get(code, {"speciesCode": code})
        # Inclui apenas espécies (exclui híbridos, slash, spuh)
        if info.get("category", "species") == "species":
            species.append(info)
    return species


# ---------------------------------------------------------------------------
# Etapa 5 — Persistência no MongoDB
# ---------------------------------------------------------------------------


def get_collection(db_name: str, collection_name: str):
    client = MongoClient(MONGO_URI)
    db = client[db_name]
    return db[collection_name]


def save_municipio(col, municipio: dict, species: list[dict]) -> None:
    """
    Upsert de um documento de município no MongoDB.
    Documento:
    {
      geocodigo, nome, uf, bioma, ebird_code, ebird_name, match_type,
      total_especies,
      especies: [ {speciesCode, comName, sciName, order, familyComName, …} ]
    }
    """
    doc = {
        "geocodigo": municipio["geocodigo"],
        "nome": municipio["nome"],
        "uf": municipio["uf"],
        "bioma": municipio["bioma"],
        "ebird_code": municipio.get("ebird_code"),
        "ebird_name": municipio.get("ebird_name"),
        "match_type": municipio.get("match_type"),
        "total_especies": len(species),
        "especies": species,
    }
    col.update_one(
        {"geocodigo": municipio["geocodigo"]},
        {"$set": doc},
        upsert=True,
    )


def ensure_indexes(col) -> None:
    col.create_index("geocodigo", unique=True)
    col.create_index("uf")
    col.create_index("bioma")
    col.create_index("ebird_code")
    col.create_index("especies.speciesCode")
    log.info("Índices do MongoDB criados/verificados.")


def bulk_save_batch(col, batch: list[tuple[dict, list[dict]]]) -> None:
    """Persiste um lote de (municipio, species) no MongoDB via bulk_write."""
    if not batch:
        return

    ops = []
    for municipio, species in batch:
        doc = {
            "geocodigo": municipio["geocodigo"],
            "nome": municipio["nome"],
            "uf": municipio["uf"],
            "bioma": municipio["bioma"],
            "ebird_code": municipio.get("ebird_code"),
            "ebird_name": municipio.get("ebird_name"),
            "match_type": municipio.get("match_type"),
            "total_especies": len(species),
            "especies": species,
        }
        ops.append(
            UpdateOne(
                {"geocodigo": municipio["geocodigo"]},
                {"$set": doc},
                upsert=True,
            )
        )

    try:
        result = col.bulk_write(ops, ordered=False)
        log.info(
            f"  Lote persistido: {result.upserted_count} inseridos, {result.modified_count} atualizados"
        )
    except BulkWriteError as e:
        log.error(f"Erro no bulk_write do lote: {e.details}")


# ---------------------------------------------------------------------------
# Etapa 6 — Salvar relatório de municípios sem match
# ---------------------------------------------------------------------------


def save_unmatched_report(unmatched: list[dict]) -> None:
    if not unmatched:
        return
    report_path = Path("municipios_sem_match.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("Geocódigo\tNome\tUF\tBioma\n")
        for m in unmatched:
            f.write(f"{m['geocodigo']}\t{m['nome']}\t{m['uf']}\t{m['bioma']}\n")
    log.info(f"Relatório de sem-match salvo em {report_path} ({len(unmatched)} municípios)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    if not EBIRD_API_KEY:
        raise ValueError(
            "Variável de ambiente EBIRD_API_KEY não definida. "
            "Consulte .env.example para configuração."
        )

    log.info("=" * 60)
    log.info("eBird × IBGE — início da coleta")
    log.info("=" * 60)

    # 1. Ler IBGE
    ibge = load_ibge(IBGE_XLSX)

    # 2. Cruzar nomes
    matched, unmatched = crossmatch(ibge)
    save_unmatched_report(unmatched)

    # 3. Taxonomia (cache)
    taxonomy = fetch_taxonomy()

    # 4 & 5. Coletar espécies e persistir em lotes de 500
    col = get_collection(MONGO_DB, "ebird")
    ensure_indexes(col)

    # Checkpoint: carrega geocódigos já coletados para retomar de onde parou
    log.info("Verificando municípios já coletados no MongoDB…")
    ja_coletados = {
        doc["geocodigo"]
        for doc in col.find(
            {
                "total_especies": {"$gt": 0}
            },  # considera apenas os que realmente tiveram dados coletados
            {"geocodigo": 1, "_id": 0},
        )
    }
    pendentes = [m for m in matched if m["geocodigo"] not in ja_coletados]
    log.info(f"  Já coletados: {len(ja_coletados)} | Pendentes: {len(pendentes)}")

    if not pendentes:
        log.info("Todos os municípios já foram coletados. Nada a fazer.")
    else:
        BATCH_SIZE = 500
        batch: list[tuple[dict, list[dict]]] = []
        total_processados = 0

        log.info(
            f"Iniciando coleta de espécies para {len(pendentes)} municípios pendentes (lotes de {BATCH_SIZE})…"
        )
        for mun in tqdm(pendentes, desc="Municípios"):
            species = fetch_species_for_municipio(mun["ebird_code"], taxonomy)
            batch.append((mun, species))

            if len(batch) >= BATCH_SIZE:
                tqdm.write(
                    f"[INFO] Persistindo lote ({total_processados + 1}–{total_processados + len(batch)})…"
                )
                log.info(
                    f"Persistindo lote ({total_processados + 1}–{total_processados + len(batch)})…"
                )
                bulk_save_batch(col, batch)
                total_processados += len(batch)
                batch = []

        # Persiste o lote final (menor que BATCH_SIZE)
        if batch:
            log.info(
                f"Persistindo lote final ({total_processados + 1}–{total_processados + len(batch)})…"
            )
            bulk_save_batch(col, batch)
            total_processados += len(batch)

        log.info(f"  Municípios processados nesta execução: {total_processados}")

    # Persistir também os sem-match (sem espécies) que ainda não existem
    sem_match_pendentes = [mun for mun in unmatched if mun["geocodigo"] not in ja_coletados]
    if sem_match_pendentes:
        log.info(f"Persistindo {len(sem_match_pendentes)} municípios sem match…")
        bulk_save_batch(col, [(mun, []) for mun in sem_match_pendentes])

    log.info("Coleta finalizada.")
    log.info(f"  Total no MongoDB: {col.count_documents({})}")
    log.info(f"  Municípios sem match eBird: {len(unmatched)}")
    log.info(f"  MongoDB: {MONGO_URI} / db={MONGO_DB} / col=municipios")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
